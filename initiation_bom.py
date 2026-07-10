#!/usr/bin/env python3
"""
Decorpot - Initiation BOM Automation  (config-driven)
=====================================================

Generates an ERP-ready Initiation BOM from a quotation CSV.

ALL business rules live in a single master Excel (default:
Initiation_BOM_Master_Data.xlsx) with three sheets:

  1. "Quantity Mapping"  - ERP CSV line item(s) -> Initiation BOM item + logic
  2. "Attribute Master"  - Initiation BOM item -> Scope / Sub-category / UOM
  3. "Settings"          - manual-review, manual-name, merge, exclude, factor,
                            false-ceiling routing

The team maintains that ONE file; no code changes are needed to add items,
change attributes, or adjust behaviour.

Usage:
  python3 initiation_bom.py <quotation.csv> <master.xlsx> -o <output.xlsx>
"""

import csv
import os
import re
import argparse
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BRAND_DARK = "303030"
BRAND_ORANGE = "F26A30"
MANUAL_HL = "FFF3E0"
DEFAULT_ATTR = ("Services", "Electrical", "NO")


def norm(s):
    if s is None:
        return ""
    s = str(s).replace("\t", " ")
    return re.sub(r"\s+", " ", s).strip().lower()


def clean_bom_name(name):
    s = str(name).replace("\t", " ").strip()
    parts = [p.strip() for p in s.split("\n") if p.strip()]
    if len(parts) <= 1:
        return re.sub(r"\s+", " ", s).strip()
    first = parts[0]
    if "_" in first:
        return first.split("_")[0].strip()
    return re.sub(r"\s+", " ", first).strip()


def parse_logic(logic_text, floor_factor):
    t = (logic_text or "").strip().lower()
    m = re.search(r"1\s*qty\s*=\s*(\d+(?:\.\d+)?)\s*sq", t)
    if m:
        return ("qty", float(m.group(1)))
    if "area" in t:
        return ("area", 1.0)
    return ("qty", 1.0)


def to_number(val):
    if val is None:
        return 0.0
    s = str(val).replace(",", "").strip()
    if s == "":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def fmt_num(x):
    if isinstance(x, str):
        return x
    return int(x) if x == int(x) else round(x, 2)


def load_config(master_path):
    wb = openpyxl.load_workbook(master_path, data_only=True)

    def sheet(*names):
        for n in names:
            if n in wb.sheetnames:
                return wb[n]
        return None

    cfg = {
        "attr": {}, "manual_review": set(), "manual_name": set(),
        "merge": set(), "exclude": set(), "floor_factor": 70.0,
        "fc_routing": True, "mapping": defaultdict(list),
        "fixed_qty": {}, "consolidate": set(),
    }

    ws = sheet("Settings")
    if ws:
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            key = str(r[0]).strip().upper()
            val = str(r[1]).strip() if r[1] is not None else ""
            vlow = val.lower()
            if key == "MANUAL_REVIEW" and val:
                cfg["manual_review"].add(vlow)
            elif key == "MANUAL_NAME" and val:
                cfg["manual_name"].add(vlow)
            elif key == "MERGE_WITHIN_ROOM" and val:
                cfg["merge"].add(vlow)
            elif key == "EXCLUDE" and val:
                cfg["exclude"].add(vlow)
            elif key == "FLOOR_MATT_FACTOR" and val:
                cfg["floor_factor"] = to_number(val) or 70.0
            elif key == "FALSE_CEILING_ROUTING":
                cfg["fc_routing"] = vlow in ("on", "true", "yes", "1")
            elif key == "FIXED_QTY" and val and "=" in val:
                item, num = val.rsplit("=", 1)
                cfg["fixed_qty"][item.strip().lower()] = to_number(num)
            elif key == "CONSOLIDATE_ALL_AREA" and val:
                cfg["consolidate"].add(vlow)

    ws = sheet("Attribute Master")
    if ws:
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            cfg["attr"][norm(r[0])] = (
                str(r[1] or "Services").strip(),
                str(r[2] or "Others").strip(),
                str(r[3] or "NO").strip(),
            )

    ws = sheet("Quantity Mapping", "Sheet1")
    if ws:
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0] or not str(r[1] or "").strip():
                continue
            bom = clean_bom_name(r[0])
            logic = str(r[2] or "").strip()
            mode, factor = parse_logic(logic, cfg["floor_factor"])
            for part in str(r[1]).split("\n"):
                k = norm(part)
                if k:
                    cfg["mapping"][k].append(
                        {"bom": bom.strip(), "mode": mode,
                         "factor": factor, "logic": logic})

    # False Ceiling: each ERP name is assigned to its target ("False Ceiling"
    # or "Paint for False Ceiling") directly in the Quantity Mapping sheet by
    # exact name -- no keyword logic needed.

    return cfg


def process_csv(csv_path, cfg):
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        cols = {c.lower().strip(): c for c in reader.fieldnames}

        def pick(*aliases):
            for a in aliases:
                if a in cols:
                    return cols[a]
            return None

        c_prod = pick("productname", "name")
        c_qty = pick("qty")
        c_area = pick("area", "area/qty(in sq ft)")
        c_room = pick("categoryname", "category")
        if not c_prod:
            raise SystemExit("ERROR: CSV has no product-name column.")
        rows = list(reader)

    csv_room_order = []
    for row in rows:
        rm = (row.get(c_room) or "").strip().upper() if c_room else ""
        if rm and rm not in csv_room_order:
            csv_room_order.append(rm)

    raw, unmatched = [], []
    for row in rows:
        prod = (row.get(c_prod) or "").strip()
        if not prod:
            continue
        maps = cfg["mapping"].get(norm(prod))
        if not maps:
            unmatched.append(prod)
            continue
        qty = to_number(row.get(c_qty))
        area = to_number(row.get(c_area))
        room = (row.get(c_room) or "").strip() if c_room else ""
        for mp in maps:
            bl = mp["bom"].lower().strip()
            if bl in cfg["exclude"]:
                continue
            if mp["mode"] == "area":
                value, basis = area * mp["factor"], "area"
            else:
                value, basis = qty * mp["factor"], "qty"
            if bl in cfg["fixed_qty"]:          # e.g. Gas Piping always = 1
                value = cfg["fixed_qty"][bl]
            raw.append({"room": room, "bom_item": mp["bom"], "bom_lower": bl,
                        "erp_product": prod, "value": value, "basis": basis,
                        "src_qty": qty, "src_area": area})

    merged, order, final = {}, [], []
    for r in raw:
        if r["bom_lower"] in cfg["merge"]:
            k = (r["room"], r["bom_lower"])
            if k not in merged:
                merged[k] = dict(r)
                order.append(k)
            else:
                merged[k]["value"] += r["value"]
                merged[k]["src_qty"] += r["src_qty"]
        else:
            final.append(r)
    for k in order:
        final.append(merged[k])

    # consolidate configured items across all rooms into ONE line (Room=ALL AREA)
    if cfg["consolidate"]:
        consolidated = {}
        corder = []
        kept = []
        for r in final:
            if r["bom_lower"] in cfg["consolidate"]:
                if r["bom_lower"] not in consolidated:
                    c = dict(r)
                    c["room"] = "ALL AREA"
                    consolidated[r["bom_lower"]] = c
                    corder.append(r["bom_lower"])
                else:
                    consolidated[r["bom_lower"]]["value"] += r["value"]
            else:
                kept.append(r)
        for bl in corder:
            kept.append(consolidated[bl])
        final = kept

    results = []
    for r in final:
        manual = (r["bom_lower"] in cfg["manual_review"]
                  or r["bom_lower"] in cfg["manual_name"])
        scope, subcat, uom = cfg["attr"].get(r["bom_lower"], DEFAULT_ATTR)
        results.append({
            "room": r["room"], "scope": scope, "subcat": subcat,
            "material": r["bom_item"], "uom": uom,
            "qty_value": fmt_num(r["value"]), "manual": manual})

    def rank(room):
        ru = room.upper()
        return csv_room_order.index(ru) if ru in csv_room_order else len(csv_room_order)
    results.sort(key=lambda rec: rank(rec["room"]))

    return results, unmatched


def write_output(results, unmatched, out_path, template=None):
    """
    If `template` (path to the 3-sheet ERP file: Lists / Price / BOM) is given,
    copy it verbatim and fill only the BOM sheet. Otherwise fall back to a
    standalone single-sheet workbook.
    """
    if template and os.path.exists(template):
        _write_with_template(results, out_path, template)
    else:
        _write_standalone(results, unmatched, out_path)


def _write_with_template(results, out_path, template):
    import copy
    wb = openpyxl.load_workbook(template)
    if "BOM" not in wb.sheetnames:
        return _write_standalone(results, [], out_path)
    ws = wb["BOM"]

    # capture the styling of the first existing data row (row 2) to reuse
    sample = {c: ws.cell(row=2, column=c) for c in range(1, 7)} if ws.max_row >= 2 else {}

    def styled(r, c, value, manual):
        cell = ws.cell(row=r, column=c)
        cell.value = value
        s = sample.get(c)
        if s is not None:
            cell.font = copy.copy(s.font)
            cell.border = copy.copy(s.border)
            cell.alignment = copy.copy(s.alignment)
            cell.number_format = s.number_format
        if manual:
            cell.fill = PatternFill("solid", fgColor=MANUAL_HL)
        return cell

    # clear existing data rows (keep header row 1)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # write our rows starting row 2
    r = 2
    for it in results:
        vals = [it["room"], it["scope"], it["subcat"],
                it["material"], it["uom"], it["qty_value"]]
        for c, v in enumerate(vals, start=1):
            styled(r, c, v, it["manual"])
        r += 1

    wb.save(out_path)


def _write_standalone(results, unmatched, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    headers = ["Select Room", "Select Scope", "Scope Sub category",
               "Select Material", "UOM", "Qty"]
    ws.append(headers)

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", fgColor=BRAND_DARK)
    hfont = Font(bold=True, color="FFFFFF", size=11)
    mfill = PatternFill("solid", fgColor=MANUAL_HL)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = hfill, hfont
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for it in results:
        ws.append([it["room"], it["scope"], it["subcat"],
                   it["material"], it["uom"], it["qty_value"]])

    for r in range(2, ws.max_row + 1):
        rec = results[r - 2]
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 4))
            if rec["manual"]:
                cell.fill = mfill
        ws.cell(row=r, column=6).font = Font(
            bold=True, color="E65100" if rec["manual"] else BRAND_ORANGE)

    for i, w in enumerate([16, 18, 18, 38, 8, 12], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Unmatched (not in mapping)")
    ws2.append(["ERP Product Name", "Occurrences"])
    for c in range(1, 3):
        cell = ws2.cell(row=1, column=c)
        cell.fill, cell.font = hfill, hfont
        cell.alignment = Alignment(horizontal="center")
    counts = defaultdict(int)
    for u in unmatched:
        counts[u] += 1
    for name in sorted(counts):
        ws2.append([name, counts[name]])
    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 14
    ws2.freeze_panes = "A2"

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="Decorpot Initiation BOM automation")
    ap.add_argument("csv", help="ERP quotation CSV")
    ap.add_argument("master", help="Master data Excel")
    ap.add_argument("-o", "--out", default="Initiation_BOM_Output.xlsx")
    ap.add_argument("-t", "--template", default=None,
                    help="3-sheet ERP template (Lists/Price/BOM) to fill")
    args = ap.parse_args()

    cfg = load_config(args.master)
    results, unmatched = process_csv(args.csv, cfg)
    write_output(results, unmatched, args.out, template=args.template)

    print(f"Mapping entries:            {len(cfg['mapping'])}")
    print(f"Initiation BOM lines:       {len(results)}")
    print(f"Unmatched CSV products:     {len(set(unmatched))}")
    print(f"Output:                     {args.out}")


if __name__ == "__main__":
    main()
